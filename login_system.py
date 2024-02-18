import openpyxl
from tkinter import messagebox

class status:
    session_status = False
    login_type = ""
    user_id = ""

class Login_System():#<---------Login database using ms exel
    def __init__(self, excel_file='credentials.xlsx'):
        self.excel_file = excel_file

        # Create the Excel file if it doesn't exist
        try:
            self.workbook = openpyxl.load_workbook(self.excel_file)
            self.sheet = self.workbook.active
        except FileNotFoundError:
            self.create_excel_file()

    def create_excel_file(self):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet['A1'] = 'Username'
        self.sheet['B1'] = 'Password'
        self.sheet['C1'] = "Type"
        self.workbook.save(self.excel_file)

    def authenticate(self, username, password, type):
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            stored_username, stored_password, stored_type = row
            if username == stored_username and str(password) == str(stored_password) and type == stored_type :
                return True
        return False

    # def register(self, username, password):    #only use 2 parameter
    #     self.sheet.append([username, password])
    #     self.workbook.save(self.excel_file)
    #     messagebox.showinfo("Registration Successful", "Account registered successfully!")

    def login(self, entered_username, entered_password, entered_type):
        if self.authenticate(entered_username, entered_password, entered_type) and entered_type == "client":
            messagebox.showinfo("Login Successful", "Welcome, {}".format(entered_username))
            status.login_type = "client"
            status.session_status = True
            status.user_id = entered_username
        elif self.authenticate(entered_username, entered_password, entered_type) and entered_type == "admin":
            messagebox.showinfo("Login Successful", "Welcome, {}".format(entered_username))
            status.login_type = "admin"
            status.session_status = True
        else:
            messagebox.showerror("Login Failed", "Invalid username or password")
            status.login_type = ""
            status.session_status = False


    # def register_user(self, entered_username, entered_password):
    #     if entered_username and entered_password:
    #         self.register(entered_username, entered_password)
    #     else:
    #         messagebox.showerror("Registration Failed", "Username and password are required.")