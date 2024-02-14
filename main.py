import tkinter as tk
from tkinter import messagebox
import openpyxl
from pathlib import Path
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage
import time
import socket


class status:
    session_status = False
    login_type = ""
    user_id = ""

    def get_ip_address():
        try:
            hostname = socket.gethostname()
            ip_address = socket.gethostbyname(hostname)
            return ip_address
        except Exception as e:
            print("Error:", e)
            return None
    ip_address = get_ip_address()
    IP = ip_address

class Login_System(status):#<---------Login database using ms exel
    def __init__(self, excel_file='credentials.xlsx'):
        self.excel_file = excel_file

        # Create the Excel file if it doesn't exist
        try:
            self.workbook = openpyxl.load_workbook(self.excel_file)
            self.sheet = self.workbook.active
        except FileNotFoundError:
            self.create_excel_file()

    def create_excel_file(self):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet['A1'] = 'Username'
        self.sheet['B1'] = 'Password'
        self.sheet['C1'] = "Type"
        self.workbook.save(self.excel_file)

    def authenticate(self, username, password, type):
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            stored_username, stored_password, stored_type = row
            if username == stored_username and str(password) == str(stored_password) and type == stored_type :
                return True
        return False

    # def register(self, username, password):    #only use 2 parameter
    #     self.sheet.append([username, password])
    #     self.workbook.save(self.excel_file)
    #     messagebox.showinfo("Registration Successful", "Account registered successfully!")

    def login(self, entered_username, entered_password, entered_type):
        if self.authenticate(entered_username, entered_password, entered_type) and entered_type == "client":
            messagebox.showinfo("Login Successful", "Welcome, {}".format(entered_username))
            status.login_type = "client"
            status.session_status = True
            status.user_id = entered_username
        elif self.authenticate(entered_username, entered_password, entered_type) and entered_type == "admin":
            messagebox.showinfo("Login Successful", "Welcome, {}".format(entered_username))
            status.login_type = "admin"
            status.session_status = True
        else:
            messagebox.showerror("Login Failed", "Invalid username or password")
            status.login_type = ""
            status.session_status = False


    # def register_user(self, entered_username, entered_password):
    #     if entered_username and entered_password:
    #         self.register(entered_username, entered_password)
    #     else:
    #         messagebox.showerror("Registration Failed", "Username and password are required.")

#------------------------------------------------------------------------------
class login(Login_System):#<---------Login GUI
    def __init__(self):#<-----Client login
        super().__init__()
        self.window_client_login = tk.Tk()
        self.window_client_login.title("Client")
        self.window_client_login.geometry("744x457")
        self.window_client_login.configure(bg = "#FFFFFF")

        canvas = Canvas(
            self.window_client_login,
            bg = "#FFFFFF",
            height = 457,
            width = 744,
            bd = 0,
            highlightthickness = 0,
            relief = "ridge"
        )
        canvas.place(x = 0, y = 0)
        image_image_1 = PhotoImage(
            file=("assets/login_bg.png"))
        image_1 = canvas.create_image(
            189.0,
            228.0,
            image=image_image_1
        )

        image_image_2 = PhotoImage(
            file=("assets/kiit.png"))
        image_2 = canvas.create_image(
            189.0,
            58.0,
            image=image_image_2
        )

        canvas.create_text(
            140.0,
            120.0,
            anchor="nw",
            text="Lab Track",
            fill="#000000",
            font=("Inter Medium", 20 * -1)
        )

        canvas.create_text(
            67.0,
            167.0,
            anchor="nw",
            text="A computer lab monitoring tool",
            fill="#000000",
            font=("Inter", 16 * -1)
        )

        button_image_1 = PhotoImage(
                file=("assets/login_as_admin_button.png"))
        button_1 = Button(
            image=button_image_1,
            borderwidth=0,
            highlightthickness=0,
            command=self.login_kill_client,
            relief="flat"
        )
        button_1.place(
            x=75.0,
            y=398.0,
            width=213.0,
            height=35.0
        )

        canvas.create_text(
            495.0,
            48.0,
            anchor="nw",
            text="Student Login",
            fill="#000000",
            font=("Inter Medium", 20 * -1)
        )

        image_image_3 = PhotoImage(
            file=("assets/dp.png"))
        image_3 = canvas.create_image(
            563.0,
            129.0,
            image=image_image_3
        )

        entry_image_1 = PhotoImage(
            file=("assets/entry_1.png"))
        entry_bg_1 = canvas.create_image(
            561.5,
            298.0,
            image=entry_image_1
        )
        password_entry = Entry(
            bd=0,
            bg="#FFFFFF",
            fg="#000716",
            highlightthickness=0
        )
        password_entry.place(
            x=443.0,
            y=274.0,
            width=237.0,
            height=46.0
        )

        entry_image_2 = PhotoImage(
            file=("assets/entry_2.png"))
        entry_bg_2 = canvas.create_image(
            561.5,
            218.0,
            image=entry_image_2
        )
        username_entry = Entry(
            bd=0,
            bg="#FFFFFF",
            fg="#000716",
            highlightthickness=0
        )
        username_entry.place(
            x=443.0,
            y=194.0,
            width=237.0,
            height=46.0
        )

        button_image_2 = PhotoImage(
            file=("assets/login_button.png"))
        button_2 = Button(
            image=button_image_2,
            borderwidth=0,
            highlightthickness=0,
            command=lambda: self.login(username_entry.get(), password_entry.get(), "client"),
            relief="flat"
        )
        button_2.place(
        x=517.0,
        y=354.0,
        width=101.0,
        height=38.0
        )

        self.window_client_login.resizable(False, False)
        self.window_client_login.mainloop()
#-----------------------------------------------------------------------------------------
    def login_admin(self):#<-------------------admin login
        self.window_admin_login = tk.Tk()
        self.window_admin_login.title("Admin")
        self.window_admin_login.geometry("496x433")
        self.window_admin_login.configure(bg = "#FFFFFF")

        canvas = Canvas(
            self.window_admin_login,
            bg = "#FFFFFF",
            height = 433,
            width = 496,
            bd = 0,
            highlightthickness = 0,
            relief = "ridge"
        )

        canvas.place(x = 0, y = 0)
        image_image_1 = PhotoImage(
            file=("assets/kiit.png"))
        image_1 = canvas.create_image(
            248.0,
            52.0,
            image=image_image_1
        )

        canvas.create_text(
            172.0,
            113.0,
            anchor="nw",
            text="Admin Login",
            fill="#000000",
            font=("Inter Medium", 24 * -1)
        )

        entry_image_1 = PhotoImage(
            file=("assets/entry_1.png"))
        entry_bg_1 = canvas.create_image(
            248.5,
            191.0,
            image=entry_image_1
        )
        username_entry = Entry(
            bd=0,
            bg="#FFFFFF",
            fg="#000716",
            highlightthickness=0
        )
        username_entry.place(
            x=144.0,
            y=172.0,
            width=209.0,
            height=36.0
        )

        entry_image_2 = PhotoImage(
            file=("assets/entry_2.png"))
        entry_bg_2 = canvas.create_image(
            248.5,
            252.0,
            image=entry_image_2
        )
        password_entry = Entry(
            bd=0,
            bg="#FFFFFF",
            fg="#000716",
            highlightthickness=0
        )
        password_entry.place(
            x=144.0,
            y=233.0,
            width=209.0,
            height=36.0
        )

        button_image_1 = PhotoImage(
            file=("assets/admin_login_button.png"))
        button_1 = Button(
            image=button_image_1,
            borderwidth=0,
            highlightthickness=0,
            command=lambda: self.login(username_entry.get(), password_entry.get(), "admin"),
            relief="flat"
        )
        button_1.place(
            x=207.0,
            y=314.0,
            width=82.0,
            height=38.0
        )
        self.window_admin_login.resizable(False, False)
        self.window_admin_login.mainloop()

        # username_label = tk.Label(window_admin_login, text="Username:")
        # username_label.grid(row=0, column=0, sticky="e", pady=5)
        # username_entry = tk.Entry(window_admin_login)
        # username_entry.grid(row=0, column=1, pady=5)

        # password_label = tk.Label(window_admin_login, text="Password:")
        # password_label.grid(row=1, column=0, sticky="e", pady=5)
        # password_entry = tk.Entry(window_admin_login, show="*")
        # password_entry.grid(row=1, column=1, pady=5)

        # login_button = tk.Button(window_admin_login, text="Login", command=lambda: self.login(username_entry.get(), password_entry.get()))
        # login_button.grid(row=2, column=0, pady=10)

        # register_button = tk.Button(window_admin_login, text="Register", command=lambda: self.register_user(username_entry.get(), password_entry.get()))
        # register_button.grid(row=2, column=1, pady=10)

        # window_admin_login.mainloop()
        #---------------------------------------------------------
        # username_label = tk.Label(window_client_login, text="Username:")
        # username_label.grid(row=0, column=0, sticky="e", pady=5)
        # username_entry = tk.Entry(window_client_login)
        # username_entry.grid(row=0, column=1, pady=5)

        # password_label = tk.Label(window_client_login, text="Password:")
        # password_label.grid(row=1, column=0, sticky="e", pady=5)
        # password_entry = tk.Entry(window_client_login, show="*")
        # password_entry.grid(row=1, column=1, pady=5)

        # login_button = tk.Button(window_client_login, text="Login", command=lambda: self.login(username_entry.get(), password_entry.get()))
        # login_button.grid(row=2, column=0, pady=10)

        # register_button = tk.Button(window_client_login, text="Register", command=lambda: self.register_user(username_entry.get(), password_entry.get()))
        # register_button.grid(row=2, column=1, pady=10)

        # window_client_login.mainloop()
        # window_client_login.resizable(False, False)
        # window_client_login.mainloop()

    def login_kill_client(self):#<---kill login window && open admin login window
        self.window_client_login.destroy()
        self.login_admin()

#-------------------------------------------------------
login_instance = login()#<------------------------------Start here
#-------------------------------------------------------
class Admin_Main():#<---------main admin window
    def __init__(self):
        super().__init__()
        self.window = tk.Tk()
        self.window.title("Client")
        self.window.geometry("744x539")
        self.window.configure(bg = "#A5938D")

        canvas = Canvas(
            self.window,
            bg = "#A5938D",
            height = 539,
            width = 744,
            bd = 0,
            highlightthickness = 0,
            relief = "ridge"
        )

        canvas.place(x = 0, y = 0)
        image_image_1 = PhotoImage(
            file=("assets/kiit.png"))
        image_1 = canvas.create_image(
            372.0,
            52.0,
            image=image_image_1
        )

        canvas.create_rectangle(
            27.0,
            90.0,
            269.0,
            236.0,
            fill="#D9D9D9",
            outline="")

        button_image_1 = PhotoImage(
            file=("assets/close_server_button.png"))
        button_1 = Button(
            image=button_image_1,
            borderwidth=0,
            highlightthickness=0,
            command=lambda: print("button_1 clicked"),
            relief="flat"
        )
        button_1.place(
            x=155.0,
            y=120.0,
            width=76.0,
            height=30.0
        )

        button_image_2 = PhotoImage(
            file=("assets/start_server_button.png"))
        button_2 = Button(
            image=button_image_2,
            borderwidth=0,
            highlightthickness=0,
            command=lambda: print("button_2 clicked"),
            relief="flat"
        )
        button_2.place(
            x=46.0,
            y=120.0,
            width=76.0,
            height=26.0
        )

        canvas.create_text(
            115.0,
            95.0,
            anchor="nw",
            text="Options",
            fill="#000000",
            font=("Inter", 12 * -1)
        )

        button_image_3 = PhotoImage(
            file=("assets/port_button.png"))
        button_3 = Button(
            image=button_image_3,
            borderwidth=0,
            highlightthickness=0,
            command=lambda: print("button_3 clicked"),
            relief="flat"
        )
        button_3.place(
            x=48.0,
            y=168.0,
            width=51.648651123046875,
            height=28.0
        )

        entry_image_1 = PhotoImage(
            file=("assets/port_entry.png"))
        entry_bg_1 = canvas.create_image(
            182.5,
            182.0,
            image=entry_image_1
        )
        entry_1 = Entry(
            bd=0,
            bg="#FFFFFF",
            fg="#000716",
            highlightthickness=0
        )
        entry_1.place(
            x=114.0,
            y=168.0,
            width=137.0,
            height=26.0
        )

        entry_image_2 = PhotoImage(
            file=("assets/server_console.png"))
        entry_bg_2 = canvas.create_image(
            148.0,
            381.5,
            image=entry_image_2
        )
        entry_2 = Text(
            bd=0,
            bg="#D9D9D9",
            fg="#000716",
            highlightthickness=0
        )
        entry_2.place(
            x=27.0,
            y=254.0,
            width=242.0,
            height=253.0
        )

        entry_image_3 = PhotoImage(
            file=("assets/network_user_list.png"))
        entry_bg_3 = canvas.create_image(
            521.5,
            306.0,
            image=entry_image_3
        )
        entry_3 = Text(
            bd=0,
            bg="#D9D9D9",
            fg="#000716",
            highlightthickness=0
        )
        entry_3.place(
            x=327.0,
            y=103.0,
            width=389.0,
            height=404.0
        )
        self.window.resizable(False, False)
        self.window.mainloop()

#-------------------------------------------------------
class Client_Main():#<---------main admin window
    def __init__(self):
        super().__init__()
        self.window = tk.Tk()
        self.window.title("Admin")
        self.window.geometry("744x457")
        self.window.configure(bg = "#FFFFFF")
        self.window.overrideredirect(True)  # Remove the title bar
        # Set window position to the right side of the screen
        screen_width = self.window.winfo_screenwidth()
        self.window.geometry(f"455x768+{screen_width - 455}+0")
        # Set window level to below all other windows
        self.window.attributes('-topmost', 0)
        # def lower_window(self):
        # self.window.lower()
        # self.window.bind('<FocusIn>', lower_window)

        canvas = Canvas(
            self.window,
            bg = "#FFFFFF",
            height = 768,
            width = 455,
            bd = 0,
            highlightthickness = 0,
            relief = "ridge"
        )

        canvas.place(x = 0, y = 0)
        image_image_1 = PhotoImage(
            file=("assets/client_main_bg.png"))
        image_1 = canvas.create_image(
            341.0,
            384.0,
            image=image_image_1
        )

        image_image_2 = PhotoImage(
            file=("assets/kiit.png"))
        image_2 = canvas.create_image(
            225.0,
            68.0,
            image=image_image_2
        )

        canvas.create_text(
            148.0,
            134.0,
            anchor="nw",
            text="Lab Track",
            fill="#1E1E1E",
            font=("Inter Bold", 32 * -1)
        )

        canvas.create_rectangle(
            36.0,
            215.0,
            429.0,
            742.0,
            fill="#8D8D8D",
            outline="")

        canvas.create_rectangle(
            58.0,
            473.0,
            407.0,
            721.0,
            fill="#D9D9D9",
            outline="")

        canvas.create_rectangle(
            58.0,
            234.0,
            407.0,
            339.0,
            fill="#DEDADA",
            outline="")

        canvas.create_text(
            183.0,
            298.0,
            anchor="nw",
            text="Null3",#<---student name
            fill="#000000",
            font=("Inter Medium", 16 * -1)
        )

        canvas.create_text(
            183.0,
            273.0,
            anchor="nw",
            text=status.user_id,
            fill="#000000",
            font=("Inter Medium", 16 * -1)
        )

        canvas.create_text(
            61.0,
            298.0,
            anchor="nw",
            text="Student name:",
            fill="#000000",
            font=("Inter Medium", 16 * -1)
        )

        canvas.create_text(
            78.0,
            274.0,
            anchor="nw",
            text="  Student ID:",
            fill="#000000",
            font=("Inter Medium", 16 * -1)
        )

        canvas.create_text(
            71.0,
            248.0,
            anchor="nw",
            text="Computer ID:",
            fill="#000000",
            font=("Inter Medium", 16 * -1)
        )

        canvas.create_text(
            183.0,
            248.0,
            anchor="nw",
            text=status.IP,
            fill="#000000",
            font=("Inter Medium", 16 * -1)
        )

        canvas.create_rectangle(
            58.0,
            358.0,
            407.0,
            454.0,
            fill="#D9D9D9",
            outline="")

        canvas.create_text(
            183.0,
            404.0,
            anchor="nw",
            text="Null4",
            fill="#000000",
            font=("Inter Medium", 16 * -1)
        )

        canvas.create_text(
            66.0,
            404.0,
            anchor="nw",
            text="Server Status:",
            fill="#000000",
            font=("Inter Medium", 16 * -1)
        )

        canvas.create_text(
            183.0,
            369.0,
            anchor="nw",
            text="Null5",
            fill="#000000",
            font=("Inter Medium", 16 * -1)
        )

        canvas.create_text(
            110.0,
            369.0,
            anchor="nw",
            text="Session:",
            fill="#000000",
            font=("Inter Medium", 16 * -1)
        )

        button_image_1 = PhotoImage(
            file=("assets/client_main_endsession.png"))
        button_1 = Button(
            image=button_image_1,
            borderwidth=0,
            highlightthickness=0,
            command=lambda: self.close_client_main,
            relief="flat"
        )
        button_1.place(
            x=243.0,
            y=650.0,
            width=146.0,
            height=42.0
        )

        canvas.create_text(
            78.0,
            530.0,
            anchor="nw",
            text="Remember to shutdown the\n computer properly :>",
            fill="#000000",
            font=("Inter Medium", 20 * -1)
        )
        self.window.resizable(False, False)
        # Create a label for session time
        self.session_time_label = tk.Label(
            self.window,
            text="00:00:00",
            font=("Inter Medium", 16 * -1),
            fg="#000000"
        )
        self.session_time_label.place(x=183.0, y=369.0, anchor="nw")

        # Start the session timer
        self.session_start_time = time.time()

        self.update_session_time()
        self.window.mainloop()

    def create_canvas_elements(self):
        image_image_1 = PhotoImage(file=("assets/client_main_bg.png"))
        self.image_1 = self.canvas.create_image(341.0, 384.0, image=image_image_1)


    def update_session_time(self):
        elapsed_time = time.time() - self.session_start_time
        hours, remainder = divmod(elapsed_time, 3600)
        minutes, seconds = divmod(remainder, 60)
        time_string = "{:02d}:{:02d}:{:02d}".format(int(hours), int(minutes), int(seconds))
        self.session_time_label.config(text="Session Time: " + time_string)
        self.session_time_label.update()  # Force update of label
        self.window.after(1000, self.update_session_time)

    def close_client_main(self):
        self.window.destroy()
#-------------------------------------------------------
if(status.session_status == True and status.login_type == "admin"):
    abc = Admin_Main()
elif(status.session_status == True and status.login_type == "client"):
    abc = Client_Main()
elif(status.session_status == False):
    pass
else:
    messagebox.showerror("There was some unknown error in the code")
print(status.session_status, status.login_type)

