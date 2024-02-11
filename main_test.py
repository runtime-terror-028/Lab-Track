from pathlib import Path
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage
from tkinter import messagebox
import openpyxl

class login_system:#<---------Login database using ms exel
    def __init__(self, excel_file='credentials.xlsx'):
        self.excel_file = excel_file

        # Create the Excel file if it doesn't exist
        try:
            self.workbook = openpyxl.load_workbook(self.excel_file)
            self.sheet = self.workbook.active
        except FileNotFoundError:
            self.create_excel_file()

    def create_excel_file(self):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet['A1'] = 'Username'
        self.sheet['B1'] = 'Password'
        self.workbook.save(self.excel_file)

    def authenticate(self, username, password):
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            stored_username, stored_password = row
            if username == stored_username and password == stored_password:
                return True
        return False

    # def register(self, username, password):
    #     self.sheet.append([username, password])
    #     self.workbook.save(self.excel_file)
    #     messagebox.showinfo("Registration Successful", "Account registered successfully!")

    def login(self, entered_username, entered_password):
        if self.authenticate(entered_username, entered_password):
            messagebox.showinfo("Login Successful", "Welcome, {}".format(entered_username))
        else:
            messagebox.showerror("Login Failed", "Invalid username or password")

    # def register_user(self, entered_username, entered_password):
    #     if entered_username and entered_password:
    #         self.register(entered_username, entered_password)
    #     else:
    #         messagebox.showerror("Registration Failed", "Username and password are required.")
#------------------------------------------------------------------------------

OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH = OUTPUT_PATH / Path(r"assets")

def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)

window = Tk()
window.geometry("744x457")
window.configure(bg = "#FFFFFF")

canvas = Canvas(
    window,
    bg = "#FFFFFF",
    height = 457,
    width = 744,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge"
)

canvas.place(x = 0, y = 0)
image_image_1 = PhotoImage(
    file=relative_to_assets("login_bg.png"))
image_1 = canvas.create_image(
    189.0,
    228.0,
    image=image_image_1
)

image_image_2 = PhotoImage(
    file=relative_to_assets("kiit.png"))
image_2 = canvas.create_image(
    189.0,
    58.0,
    image=image_image_2
)

canvas.create_text(
    140.0,
    120.0,
    anchor="nw",
    text="Lab Track",
    fill="#000000",
    font=("Inter Medium", 20 * -1)
)

canvas.create_text(
    67.0,
    167.0,
    anchor="nw",
    text="A computer lab monitoring tool",
    fill="#000000",
    font=("Inter", 16 * -1)
)

button_image_1 = PhotoImage(
    file=relative_to_assets("login_as_admin_button.png"))
button_1 = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: print("button_1 clicked"),
    relief="flat"
)
button_1.place(
    x=75.0,
    y=398.0,
    width=213.0,
    height=35.0
)

canvas.create_text(
    495.0,
    48.0,
    anchor="nw",
    text="Student Login",
    fill="#000000",
    font=("Inter Medium", 20 * -1)
)

image_image_3 = PhotoImage(
    file=relative_to_assets("dp.png"))
image_3 = canvas.create_image(
    563.0,
    129.0,
    image=image_image_3
)

entry_image_1 = PhotoImage(
    file=relative_to_assets("entry_1.png"))
entry_bg_1 = canvas.create_image(
    561.5,
    298.0,
    image=entry_image_1
)
entry_1 = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#000716",
    highlightthickness=0
)
entry_1.place(
    x=443.0,
    y=274.0,
    width=237.0,
    height=46.0
)

entry_image_2 = PhotoImage(
    file=relative_to_assets("entry_2.png"))
entry_bg_2 = canvas.create_image(
    561.5,
    218.0,
    image=entry_image_2
)
entry_2 = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#000716",
    highlightthickness=0
)
entry_2.place(
    x=443.0,
    y=194.0,
    width=237.0,
    height=46.0
)

button_image_2 = PhotoImage(
    file=relative_to_assets("button_2.png"))
button_2 = Button(
    image=button_image_2,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: print("button_2 clicked"),
    relief="flat"
)
button_2.place(
    x=517.0,
    y=354.0,
    width=101.0,
    height=38.0
)
window.resizable(False, False)
window.mainloop()

